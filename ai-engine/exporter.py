"""
Excel Exporter

Exports QuantityResult to a formatted Excel (.xlsx) file.
"""

from pathlib import Path

from config import EXCEL_OUTPUT
from models import QuantityResult
from utils  import print_success, print_error


class Exporter:

    def export(
        self,
        result: QuantityResult,
        output_path: Path = EXCEL_OUTPUT,
    ) -> Path:
        """
        Write quantities to Excel.
        Requires openpyxl (add to requirements.txt if missing).
        """

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print_error(
                "openpyxl not installed. "
                "Run: pip install openpyxl"
            )
            raise

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quantities"

        # ─── Header ──────────────────────────────────────────
        header_fill = PatternFill(
            start_color="1F3864",
            end_color="1F3864",
            fill_type="solid",
        )
        header_font = Font(
            color="FFFFFF",
            bold=True,
            size=12,
        )

        headers = ["#", "Description", "Quantity", "Unit"]
        col_widths = [5, 40, 15, 10]

        for col_idx, (header, width) in enumerate(
            zip(headers, col_widths), start=1
        ):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = width

        # ─── Drawing type row ─────────────────────────────────
        ws.cell(row=2, column=1, value="Type")
        ws.cell(row=2, column=2, value=result.drawing_type.value.capitalize())
        ws.cell(row=2, column=2).font = Font(italic=True, color="555555")

        # ─── Data rows ───────────────────────────────────────
        alt_fill = PatternFill(
            start_color="EEF2F7",
            end_color="EEF2F7",
            fill_type="solid",
        )

        for row_idx, item in enumerate(result.items, start=3):
            ws.cell(row=row_idx, column=1, value=row_idx - 2)
            ws.cell(row=row_idx, column=2, value=item.name)
            ws.cell(row=row_idx, column=3, value=item.quantity)
            ws.cell(row=row_idx, column=4, value=item.unit)

            # Alternate row shading
            if row_idx % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = alt_fill

        wb.save(output_path)

        print_success(f"Excel exported → {output_path}")

        return output_path


if __name__ == "__main__":

    from config import QUANTITY_OUTPUT
    from utils  import load_json

    data   = load_json(QUANTITY_OUTPUT)
    result = QuantityResult(**data)

    exporter = Exporter()
    exporter.export(result)
